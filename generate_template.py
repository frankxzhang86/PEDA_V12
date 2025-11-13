import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MANDATORY_FIELDS = {"part_number", "reason", "decision_region", "decision_value"}


def autofit_columns(sheet, minimum=12, maximum=60, padding=2):
    """Auto-adjust column widths based on the longest cell content."""
    for column_cells in sheet.columns:
        first_cell = column_cells[0]
        column_letter = getattr(first_cell, "column_letter", None)
        if column_letter is None:
            column_letter = get_column_letter(first_cell.column)

        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length

        adjusted_width = max(minimum, min(max_length + padding, maximum))
        sheet.column_dimensions[column_letter].width = adjusted_width


def style_instruction_sheet(sheet):
    """Apply light styling to the Instructions sheet."""
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center_align = Alignment(vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    sheet.freeze_panes = "A2"
    autofit_columns(sheet, minimum=15, maximum=90)


def style_data_sheet(sheet):
    """Highlight mandatory headers and apply consistent styling."""
    mandatory_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    optional_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        is_mandatory = str(cell.value) in MANDATORY_FIELDS
        cell.fill = mandatory_fill if is_mandatory else optional_fill
        cell.font = header_font
        cell.alignment = header_alignment

    sheet.freeze_panes = "A2"
    autofit_columns(sheet, minimum=15, maximum=50)


def create_peda_upload_template():
    """
    生成 PEDA V12 的 Excel 上传模板文件。
    该文件包含一个数据输入表和一个使用说明表。
    """
    template_filename = "PEDA_Upload_Template.xlsx"
    
    # --- 1. 创建数据模板工作表 (Data Sheet) ---
    # 根据新的字段要求：4个必填 + 5个选填
    data_template = {
        "part_number": ["PN-001-A", "PN-002-B", "PN-003-C"],
        "reason": ["250", "250", "250"],
        "decision_region": ["Asia", "Europe", "Asia"],
        "decision_value": ["10", "10", "10"],
        "contact": ["Pipar Pan", "Pipar Pan", ""],
        "external_info": ["External information 1", "External information 2", ""],
        "internal_comment": ["Internal comment 1", "Internal comment 2", ""],
        "project_type": ["2", "2", ""],
        "sample_quantity": ["10", "20", ""]
    }
    df_template = pd.DataFrame(data_template)

    # --- 2. 创建使用说明工作表 (Instruction Sheet) ---
    instructions_data = {
        "字段名 (Field Name)": [
            "part_number",
            "reason", 
            "decision_region",
            "decision_value",
            "contact",
            "external_info",
            "internal_comment",
            "project_type",
            "sample_quantity"
        ],
        "说明 (Description)": [
            "【必填】产品料号，系统会根据此料号搜索产品并创建PEDA。示例：PN-001-A",
            
            "【必填】原因代码，系统预定义的值。示例：250",
            
            "【必填】决策区域，产品适用的地区。示例：Asia, Europe",
            
            "【必填】决策值，整数。示例：10",
            
            "【选填】联系人名称，如为空则使用默认值。示例：Pipar Pan",
            
            "【选填】外部信息，可填写给外部查看的信息。如为空则留空。示例：External information",
            
            "【选填】内部备注，可填写内部使用的备注信息。如为空则留空。示例：Internal comment",
            
            "【选填】项目类型，如为空则使用默认值。示例：2",
            
            "【选填】样品数量，如为空则使用默认值。示例：10"
        ],
        "示例值 (Example)": [
            "PN-001-A",
            "250",
            "Asia", 
            "10",
            "Pipar Pan",
            "External information",
            "Internal comment",
            "2",
            "10"
        ]
    }
    df_instructions = pd.DataFrame(instructions_data)

    # --- 3. 将两个 DataFrame 写入同一个 Excel 文件 ---
    try:
        with pd.ExcelWriter(template_filename, engine='openpyxl') as writer:
            # 写入数据模板工作表
            df_template.to_excel(writer, sheet_name='PEDA Upload Data', index=False)
            
            # 写入使用说明工作表
            df_instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            # 获取工作簿对象以便进行格式调整
            workbook = writer.book
            instructions_sheet = workbook['Instructions']
            data_sheet = workbook['PEDA Upload Data']

            style_instruction_sheet(instructions_sheet)
            style_data_sheet(data_sheet)
        
        print("=" * 60)
        print("✅ 成功！PEDA V12 上传模板文件已创建")
        print("=" * 60)
        print(f"📄 文件名：{template_filename}")
        print(f"📁 完整路径：{os.path.abspath(template_filename)}")
        print("\n📊 文件包含以下工作表：")
        print("  1. PEDA Upload Data - 数据输入工作表（包含3行示例数据）")
        print("  2. Instructions - 使用说明工作表（详细字段说明）")
        print("\n📋 Excel 期待的列（字段）：")
        print("  【必填字段】")
        print("  • part_number - 产品料号（必填）")
        print("  • reason - 原因代码（必填）")
        print("  • decision_region - 决策区域（必填）")
        print("  • decision_value - 决策值（必填）")
        print("\n  【选填字段】（如为空将使用默认值或留空）")
        print("  • contact - 联系人（默认值：Pipar Pan）")
        print("  • external_info - 外部信息（默认值：空）")
        print("  • internal_comment - 内部备注（默认值：空）")
        print("  • project_type - 项目类型（默认值：2）")
        print("  • sample_quantity - 样品数量（默认值：10）")
        print("\n💡 提示：")
        print("  • 请在 'PEDA Upload Data' 工作表中填写您的数据")
        print("  • 文档主目录路径请在GUI主页设置，不再从Excel读取")
        print("=" * 60)
        
    except Exception as e:
        print("=" * 60)
        print("❌ 创建模板文件失败")
        print("=" * 60)
        print(f"错误信息：{e}")
        print("\n请确保：")
        print("  1. 已安装必要的库：pip install pandas openpyxl")
        print("  2. 当前目录有写入权限")
        print("  3. 文件未被其他程序打开")
        print("  4. PEDA_Upload_Template.xlsx 文件不存在或未被占用")
        print("=" * 60)

if __name__ == "__main__":
    create_peda_upload_template()
